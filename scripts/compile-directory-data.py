#!/usr/bin/env python3
"""Compile the research workbook into the JSON snapshots the site reads.

The directories are read-only reference data that change a few times a year,
so they ship as compiled JSON rather than database rows: no migration, no
query per request. Re-run this whenever a new workbook lands.

    python3 scripts/compile-directory-data.py <workbook.xlsx>

Writes lib/internships/organisations.json and lib/research/professors.json.
'featured.json' is left alone — the tier sheets are curated by hand and this
script has never been the source of truth for them.
"""
import json
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
PLACEHOLDERS = {"", "n/a", "na", "none", "-", "—"}


def clean(value):
    """The sheets spell 'no data found' as N/A; the site wants a real null."""
    if value is None:
        return None
    text = str(value).strip()
    return None if text.lower() in PLACEHOLDERS else text


def rows_of(workbook, sheet):
    rows = list(workbook[sheet].iter_rows(values_only=True))
    header = rows[0]
    return [dict(zip(header, r)) for r in rows[1:] if any(c is not None for c in r)]


def compile_organisations(workbook):
    out = []
    for i, r in enumerate(rows_of(workbook, "Orgs by Region"), start=1):
        name = clean(r.get("Organization"))
        if not name:
            continue
        out.append({
            "id": i,
            "region": clean(r.get("Region")) or "Unknown",
            "name": name,
            "category": clean(r.get("Category")) or "Other",
            "city": clean(r.get("City")),
            "address": clean(r.get("Address")),
            "phone": clean(r.get("Phone")),
            "email": clean(r.get("Email")),
            "sourceUrl": clean(r.get("Source URL")),
        })
    return out


def compile_professors(workbook):
    out = []
    for i, r in enumerate(rows_of(workbook, "Professors"), start=1):
        name = clean(r.get("Name"))
        if not name:
            continue
        out.append({
            "id": i,
            "tier": clean(r.get("Tier")) or "Unranked",
            "university": clean(r.get("University")) or "Unknown",
            "name": name,
            "title": clean(r.get("Title")),
            "department": clean(r.get("Department")),
            "city": clean(r.get("City")),
            "email": clean(r.get("Email")),
            "phone": clean(r.get("Phone")),
            "contactType": clean(r.get("Contact Type")) or "Unknown",
            "sourceUrl": clean(r.get("Source URL")),
        })
    return out


def write(path: Path, data):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=1) + "\n", encoding="utf-8")
    print(f"  {path.relative_to(ROOT)}  ({len(data)} records)")


def main():
    if len(sys.argv) != 2:
        sys.exit(__doc__)
    book = Path(sys.argv[1])
    wb = openpyxl.load_workbook(book, read_only=True, data_only=True)
    print(f"Reading {book.name}")
    write(ROOT / "lib/internships/organisations.json", compile_organisations(wb))
    if "Professors" in wb.sheetnames:
        write(ROOT / "lib/research/professors.json", compile_professors(wb))
    wb.close()


if __name__ == "__main__":
    main()
